#!/usr/bin/env python3
"""
============================================================
MONEY FLOW SCREENER v3.0 — INSTITUTIONAL FLOW EDITION
============================================================
pip install yfinance numpy openpyxl
python screener.py
============================================================
UPGRADES FROM v2.0:
  - Chaikin Money Flow (CMF, 20-period) — true volume-direction signal
  - Money Flow Index (MFI, 14-period) — volume-weighted RSI
  - OBV slope + bullish/bearish divergence detection
  - Up/Down volume ratio — buying vs selling pressure split
  - FINRA daily short sale volume (off-exchange short pressure)
  - Short interest from yfinance (days-to-cover, float %)
  - Cross-sectional RS percentile rank across 300-stock universe (15 sectors × 20)
  - SQLite persistence: score history + rank trajectory signal
  - Scoring rebuilt: Flow bucket = 30pts (replaces crude vol/A-D)
  - CMF chart panel added per symbol
  - Full interpretation guide embedded in HTML output
============================================================
"""

import sys, os, json, math, webbrowser, sqlite3, time
import urllib.request, csv, io
from datetime import datetime, timedelta

try:
    import yfinance as yf
except ImportError:
    print("\n  pip install yfinance numpy openpyxl\n"); sys.exit(1)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("\n  pip install openpyxl\n"); sys.exit(1)

import numpy as np

# ============================================================
# SAFE JSON
# ============================================================
class SafeEncoder(json.JSONEncoder):
    def default(s, o):
        if isinstance(o, (np.integer,)): return int(o)
        if isinstance(o, (np.floating,)):
            if math.isnan(o) or math.isinf(o): return None
            return float(o)
        if isinstance(o, np.ndarray): return o.tolist()
        return super().default(o)

def safe_json(o):
    return json.dumps(o, cls=SafeEncoder).replace("NaN","null").replace("Infinity","null").replace("-Infinity","null")

# ============================================================
# CONFIG
# ============================================================
BENCHMARK = "SPY"
MACRO_TICKERS = {"VIX": "^VIX", "TNX": "^TNX", "DXY": "DX-Y.NYB"}
DB_PATH = "money_flow_history.db"

SECTOR_ETFS = [
    ("XLK","Technology","#6366f1"),("XLF","Financials","#22d3ee"),
    ("XLE","Energy","#f97316"),("XLV","Healthcare","#ec4899"),
    ("XLI","Industrials","#a78bfa"),("XLC","Comm. Services","#34d399"),
    ("XLY","Consumer Disc.","#fbbf24"),("XLP","Consumer Staples","#94a3b8"),
    ("XLU","Utilities","#fb923c"),("XLRE","Real Estate","#c084fc"),
    ("XLB","Materials","#2dd4bf"),
    ("XBI","Biotech","#f472b6"),("KRE","Reg. Banks","#67e8f9"),
    ("SOXX","Semiconductors","#818cf8"),("XME","Metals/Mining","#fcd34d"),
]
SECTOR_STOCKS = {
    "XLK":["AAPL","MSFT","NVDA","AVGO","CRM","ADBE","AMD","ORCL","CSCO","INTC",
           "NOW","INTU","PANW","MU","QCOM","TXN","KLAC","SNPS","CDNS","AMAT"],
    "XLF":["JPM","V","MA","BAC","WFC","GS","MS","AXP","C","BLK",
           "SCHW","USB","PNC","TFC","COF","SPGI","MCO","ICE","CB","MET"],
    "XLE":["XOM","CVX","COP","SLB","EOG","MPC","PSX","VLO","OXY","HAL",
           "DVN","FANG","MRO","BKR","HES","WMB","CTRA","OVV","APA","KMI"],
    "XLV":["UNH","JNJ","LLY","PFE","ABBV","MRK","TMO","ABT","DHR","BMY",
           "CVS","CI","HUM","ELV","ISRG","MDT","BSX","SYK","ZBH","BAX"],
    "XLI":["GE","CAT","HON","UNP","RTX","DE","BA","LMT","FDX","WM",
           "EMR","ITW","ETN","PCAR","CTAS","GD","NOC","NSC","CSX","MMM"],
    "XLC":["META","GOOGL","NFLX","DIS","CMCSA","T","VZ","TMUS","EA","TTWO",
           "WBD","PARA","FOXA","LYV","OMC","IPG","ZM","SNAP","PINS","NWSA"],
    "XLY":["AMZN","TSLA","HD","MCD","NKE","LOW","SBUX","TJX","BKNG","CMG",
           "GM","F","RCL","CCL","MGM","YUM","DRI","ULTA","GPC","ROST"],
    "XLP":["PG","KO","PEP","COST","WMT","PM","MO","CL","MDLZ","GIS",
           "STZ","HSY","CPB","CAG","KHC","CHD","CLX","SJM","MKC","BG"],
    "XLU":["NEE","SO","DUK","SRE","AEP","D","EXC","XEL","ED","WEC",
           "ES","AWK","CMS","PEG","PPL","FE","NI","EVRG","AEE","LNT"],
    "XLRE":["PLD","AMT","CCI","EQIX","PSA","SPG","O","WELL","DLR","AVB",
            "EQR","ESS","MAA","UDR","EXR","CUBE","VICI","GLPI","WPC","NNN"],
    "XLB":["LIN","APD","SHW","ECL","FCX","NEM","NUE","DOW","DD","VMC",
           "MLM","PKG","IP","IFF","CE","ALB","EMN","OLN","RPM","CCK"],
    # ---- NEW SECTORS ----
    "XBI":["MRNA","REGN","VRTX","BIIB","ILMN","BMRN","ALNY","INCY","NBIX","SRPT",
           "RARE","HALO","IONS","EXAS","ACAD","CRSP","RXRX","PCVX","TGTX","ARWR"],
    "KRE":["KEY","RF","CFG","HBAN","MTB","ZION","CMA","FITB","FHN","WAL",
           "WTFC","SNV","GBCI","BOKF","FFIN","CVBF","IBTX","FNB","UMBF","BOH"],
    "SOXX":["LRCX","MRVL","ON","MPWR","SWKS","QRVO","MCHP","ADI","NXPI","WOLF",
            "ENTG","MKSI","ONTO","ACLS","FORM","CAMT","ASML","RMBS","SLAB","ALGM"],
    "XME":["AA","CLF","MP","X","CMC","RS","STLD","HCC","HL","PAAS",
           "KGC","WPM","FNV","RGLD","GOLD","AEM","AGI","MAG","SILV","TECK"],
}

# ============================================================
# HELPERS
# ============================================================
def sf(x):
    if x is None: return None
    try:
        f = float(x)
        return None if math.isnan(f) or math.isinf(f) else round(f, 4)
    except (TypeError, ValueError):
        return None

def fmt(n, d=2): return f"{n:.{d}f}" if n is not None else "-"
def fmt_pct(n):
    if n is None: return "-"
    return f"{'+' if n>=0 else ''}{n:.2f}%"
def pc(n): return "muted" if n is None else "green" if n>=0 else "red"

# ============================================================
# TECHNICALS — ORIGINAL (preserved)
# ============================================================
def calc_sma_series(c, p):
    out = [None]*len(c)
    for i in range(p-1, len(c)): out[i] = sf(np.mean(c[i-p+1:i+1]))
    return out

def calc_rsi_series(c, p=14):
    out = [None]*len(c)
    for i in range(p, len(c)):
        d = np.diff(c[i-p:i+1])
        g = np.where(d>0,d,0); l = np.where(d<0,-d,0)
        ag, al = float(np.mean(g)), float(np.mean(l))
        out[i] = 100.0 if al==0 else round(100-100/(1+ag/al),1)
    return out

def calc_sma(c, p):
    if len(c)<p: return None
    return sf(np.mean(c[-p:]))

def calc_rsi(c, p=14):
    if len(c)<p+1: return None
    d = np.diff(c[-(p+1):])
    g = np.where(d>0,d,0); l = np.where(d<0,-d,0)
    ag, al = float(np.mean(g)), float(np.mean(l))
    return 100.0 if al==0 else round(100-100/(1+ag/al),1)

def calc_atr(h, l, c, p=14):
    if len(c)<p+1: return None
    trs = [max(h[i]-l[i],abs(h[i]-c[i-1]),abs(l[i]-c[i-1])) for i in range(1,len(c))]
    return sf(np.mean(trs[-p:]))

def calc_vol_ratio(v):
    if len(v)<50: return None
    a50 = float(np.mean(v[-50:]))
    return round(float(np.mean(v[-10:]))/a50,2) if a50>0 else None

def calc_return(c, d):
    if len(c)<d+1: return None
    s = float(c[-(d+1)])
    return round(float((c[-1]-s)/s*100),2) if s>0 else None

def calc_rs_series(closes, spy):
    mn = min(len(closes),len(spy))
    c = closes[-mn:]; s = spy[-mn:]
    out = [float(c[i]/s[i]) if s[i]>0 else None for i in range(len(c))]
    if out and out[0] and out[0]>0:
        b = out[0]; out = [round(v/b*100,2) if v else None for v in out]
    return out

# ============================================================
# TECHNICALS — NEW: REAL FLOW SIGNALS
# ============================================================
def calc_cmf(closes, highs, lows, volumes, period=20):
    """
    Chaikin Money Flow: ((close-low)-(high-close))/(high-low) * volume, summed.
    Range -1 to +1. Positive = net accumulation. Negative = net distribution.
    >+0.1 = institutional buying. <-0.1 = institutional selling.
    """
    if len(closes) < period: return None
    c=closes[-period:]; h=highs[-period:]; l=lows[-period:]; v=volumes[-period:]
    mfvs = []
    for i in range(len(c)):
        rng = h[i]-l[i]
        mfvs.append(0.0 if rng==0 else ((c[i]-l[i])-(h[i]-c[i]))/rng*v[i])
    tv = sum(v)
    return round(sum(mfvs)/tv, 4) if tv>0 else None

def calc_mfi(closes, highs, lows, volumes, period=14):
    """
    Money Flow Index: RSI applied to (typical price × volume).
    55-75 = bullish sweet spot. >80 = overbought. <30 = bearish pressure.
    """
    if len(closes) < period+2: return None
    tp = [(highs[i]+lows[i]+closes[i])/3 for i in range(len(closes))]
    mf = [tp[i]*volumes[i] for i in range(len(tp))]
    pos_mf = neg_mf = 0.0
    for i in range(len(closes)-period, len(closes)):
        if i <= 0: continue
        if tp[i]>tp[i-1]: pos_mf+=mf[i]
        elif tp[i]<tp[i-1]: neg_mf+=mf[i]
    if neg_mf==0: return 100.0
    return round(100-100/(1+pos_mf/neg_mf), 1)

def calc_obv_array(closes, volumes):
    """Full OBV array."""
    obv = [0.0]
    for i in range(1, len(closes)):
        if closes[i]>closes[i-1]: obv.append(obv[-1]+volumes[i])
        elif closes[i]<closes[i-1]: obv.append(obv[-1]-volumes[i])
        else: obv.append(obv[-1])
    return obv

def calc_obv_slope(closes, volumes, period=20):
    """
    Normalized OBV linear regression slope.
    Positive = OBV trending up (accumulation). Negative = distribution.
    """
    if len(closes)<period+1: return None
    obv = calc_obv_array(closes, volumes)
    seg = np.array(obv[-period:])
    x = np.arange(len(seg))
    slope = float(np.polyfit(x, seg, 1)[0])
    avg_vol = float(np.mean(np.abs(volumes[-period:]))) or 1
    return round(slope/avg_vol, 4)

def calc_obv_divergence(closes, volumes, period=20):
    """
    Detect OBV divergence vs price over period.
    'bullish'  = price down but OBV up   → HIDDEN ACCUMULATION (strongest signal)
    'bearish'  = price up but OBV down   → DISTRIBUTION TRAP (warning)
    'confirmed_bull' = price + OBV both up → trend confirmed
    'confirmed_bear' = price + OBV both down → weakness confirmed
    'neutral'  = no significant divergence
    """
    if len(closes)<period+1: return "neutral"
    obv = calc_obv_array(closes, volumes)
    pc_ = (closes[-1]-closes[-period])/closes[-period]*100 if closes[-period]!=0 else 0
    ob0 = obv[-period] if obv[-period]!=0 else 1
    oc_ = (obv[-1]-obv[-period])/abs(ob0)*100
    pu=pc_>2; pd=pc_<-2; ou=oc_>2; od=oc_<-2
    if pu and ou: return "confirmed_bull"
    if pd and od: return "confirmed_bear"
    if pd and ou: return "bullish"
    if pu and od: return "bearish"
    return "neutral"

def calc_updown_vol(closes, volumes, period=20):
    """
    Up-day volume / down-day volume over period.
    >1.5 = buying dominates. <0.7 = selling dominates.
    """
    if len(closes)<period+1: return None
    c=closes[-period:]; v=volumes[-period:]
    up = sum(v[i] for i in range(1,len(c)) if c[i]>c[i-1])
    dn = sum(v[i] for i in range(1,len(c)) if c[i]<c[i-1])
    if dn==0: return 9.99 if up>0 else 1.0
    return round(up/dn, 2)

# ============================================================
# CROSS-SECTIONAL RS RANK
# ============================================================
def calc_rs_percentile(ret, all_rets):
    """Percentile rank of this return vs all returns in universe (0=worst, 100=best)."""
    if ret is None: return 50.0
    valid = [r for r in all_rets if r is not None]
    if not valid: return 50.0
    return round(sum(1 for r in valid if r<ret)/len(valid)*100, 1)

# ============================================================
# FINRA DAILY SHORT SALE VOLUME  (off-exchange dark pool proxy)
# ============================================================
def fetch_finra_short_volume(symbols, lookback_days=5):
    """
    Download FINRA consolidated short sale volume (CNMSshvol files).
    short_pct = ShortVolume/TotalVolume on FINRA-reported (off-exchange) trades.
    Baseline for most stocks: 40-60% (market maker hedging).
    Elevated (>55%) + rising trend = increasing bearish dark-pool pressure.
    Declining (<40%) = short covering = bullish.
    """
    print("  Fetching FINRA short sale volume (dark pool proxy)...")
    results = {s: {"short_pct": None, "short_pct_trend": None} for s in symbols}
    sym_set = set(symbols)
    sym_days = {s: [] for s in symbols}
    days_found = 0
    dt = datetime.now()
    attempts = 0

    while days_found < lookback_days and attempts < lookback_days+12:
        dt -= timedelta(days=1)
        attempts += 1
        if dt.weekday() >= 5: continue
        date_str = dt.strftime("%Y%m%d")
        url = f"https://cdn.finra.org/equity/regsho/daily/CNMSshvol{date_str}.txt"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=12) as resp:
                content = resp.read().decode("utf-8")
            reader = csv.DictReader(io.StringIO(content), delimiter="|")
            found_any = False
            for row in reader:
                sym = row.get("Symbol","").strip()
                if sym not in sym_set: continue
                try:
                    sv = float(row.get("ShortVolume",0) or 0)
                    tv = float(row.get("TotalVolume",1) or 1)
                    if tv > 0:
                        sym_days[sym].append(round(sv/tv*100, 2))
                        found_any = True
                except (ValueError, ZeroDivisionError):
                    pass
            if found_any:
                days_found += 1
                print(f"    {date_str}: loaded")
            time.sleep(0.3)
        except Exception:
            pass  # holiday or file not yet available

    for sym in symbols:
        data = sym_days.get(sym, [])
        if data:
            results[sym]["short_pct"] = data[-1]
            if len(data) >= 2:
                results[sym]["short_pct_trend"] = round(data[-1]-data[0], 2)

    found_n = sum(1 for s in symbols if results[s]["short_pct"] is not None)
    print(f"    FINRA: {found_n}/{len(symbols)} symbols with data")
    return results

# ============================================================
# SQLITE PERSISTENCE
# ============================================================
class ScoreDB:
    def __init__(self, path=DB_PATH):
        self.conn = sqlite3.connect(path)
        self._init()

    def _init(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS scores (
                date TEXT, symbol TEXT, score INTEGER,
                cmf REAL, mfi REAL, obv_slope REAL,
                updown_vol REAL, rs_pct REAL, short_pct REAL,
                PRIMARY KEY (date, symbol)
            );
            CREATE TABLE IF NOT EXISTS rank_history (
                date TEXT, symbol TEXT, rank INTEGER, universe TEXT,
                PRIMARY KEY (date, symbol, universe)
            );
        """)
        self.conn.commit()

    def save(self, date_str, symbol, d, rank=None, universe="stock"):
        try:
            self.conn.execute(
                "INSERT OR REPLACE INTO scores VALUES (?,?,?,?,?,?,?,?,?)",
                (date_str, symbol, d.get("score"), d.get("cmf"), d.get("mfi"),
                 d.get("obv_slope"), d.get("updown_vol"), d.get("rs_pct"),
                 d.get("finra_short_pct"))
            )
            if rank is not None:
                self.conn.execute(
                    "INSERT OR REPLACE INTO rank_history VALUES (?,?,?,?)",
                    (date_str, symbol, rank, universe)
                )
            self.conn.commit()
        except Exception:
            pass

    def rank_trajectory(self, symbol, universe="stock", days=5):
        """
        Negative slope = rank improving (lower rank# = better position).
        Returns (slope, [ranks oldest→newest]).
        """
        try:
            rows = self.conn.execute(
                "SELECT rank FROM rank_history WHERE symbol=? AND universe=? ORDER BY date DESC LIMIT ?",
                (symbol, universe, days)
            ).fetchall()
            ranks = [r[0] for r in reversed(rows)]
            if len(ranks) < 2: return None, ranks
            slope = float(np.polyfit(np.arange(len(ranks)), ranks, 1)[0])
            return round(slope, 2), ranks
        except Exception:
            return None, []

    def close(self):
        try: self.conn.close()
        except Exception: pass

# ============================================================
# MACRO REGIME
# ============================================================
def fetch_macro():
    print("  Fetching macro data (VIX, 10Y yield, USD)...")
    macro = {}
    for name, sym in MACRO_TICKERS.items():
        try:
            df = yf.Ticker(sym).history(period="3mo", interval="1d", auto_adjust=True)
            if not df.empty and len(df)>5:
                c = df["Close"].values.astype(float)
                macro[name] = {
                    "current": round(float(c[-1]),2),
                    "prev":    round(float(c[-2]),2) if len(c)>1 else None,
                    "wk_ago":  round(float(c[-5]),2) if len(c)>5 else None,
                    "mo_ago":  round(float(c[-21]),2) if len(c)>21 else None,
                }
                print(f"    {name}: {macro[name]['current']}")
        except Exception as e:
            print(f"    {name}: failed ({e})")
            macro[name] = {"current":None,"prev":None,"wk_ago":None,"mo_ago":None}
    return macro

def classify_regime(macro, spy_data):
    vix   = macro.get("VIX",{}).get("current")
    sma50  = calc_sma(spy_data["closes"],50)
    sma200 = calc_sma(spy_data["closes"],200)
    price  = float(spy_data["closes"][-1])
    base = {"vix":vix, "spy_price":round(price,2),
            "spy_sma50":round(sma50,2) if sma50 else None,
            "spy_sma200":round(sma200,2) if sma200 else None}
    if vix is None or sma50 is None or sma200 is None:
        return {**base,"regime":"UNKNOWN","desc":"Insufficient data","posture":"Reduce size until clarity","color":"#64748b"}
    if price>sma50 and sma50>sma200 and vix<20:
        return {**base,"regime":"TRENDING UP","desc":"SPY above rising 50 & 200 MA, VIX low. Clear uptrend.",
                "posture":"Full offense. Standard position sizes. Trust high-score signals.","color":"#22c55e"}
    elif price>sma200 and vix<28:
        return {**base,"regime":"MODERATE / CHOPPY","desc":"SPY above 200 MA but choppy around 50 MA. Mixed signals.",
                "posture":"Half positions. 70+ scores only. Confirm flow before entering.","color":"#fbbf24"}
    elif price<sma50 and price>sma200 and vix>=20:
        return {**base,"regime":"CORRECTING","desc":"SPY below 50 MA but above 200 MA. Institutional rotation underway.",
                "posture":"Quarter positions. Favor defensive sectors. Flow signals must be green.","color":"#f97316"}
    elif price<sma200 or vix>=30:
        return {**base,"regime":"RISK-OFF / BEAR","desc":"SPY below 200 MA or VIX elevated. Capital fleeing to safety.",
                "posture":"Defensive only. Consider short candidates. Preserve capital.","color":"#ef4444"}
    else:
        return {**base,"regime":"TRANSITIONAL","desc":"Market between regimes. Wait for clarity.",
                "posture":"Small positions only. Let the regime declare itself.","color":"#a78bfa"}

# ============================================================
# FUNDAMENTALS + EARNINGS + SHORT INTEREST
# ============================================================
def fetch_fundamentals(symbol):
    try:
        t = yf.Ticker(symbol)
        info = t.info or {}
        pe           = info.get("trailingPE") or info.get("forwardPE")
        rev_growth   = info.get("revenueGrowth")
        profit_margin= info.get("profitMargins")
        debt_equity  = info.get("debtToEquity")
        mkt_cap      = info.get("marketCap")
        short_ratio  = info.get("shortRatio")
        short_float  = info.get("shortPercentOfFloat")

        earnings_date = None
        try:
            cal = t.calendar
            if cal is not None:
                if isinstance(cal, dict):
                    ed = cal.get("Earnings Date")
                    if ed and len(ed)>0:
                        earnings_date = str(ed[0].date()) if hasattr(ed[0],'date') else str(ed[0])
                elif hasattr(cal,'iloc') and 'Earnings Date' in cal.index:
                    ed = cal.loc['Earnings Date']
                    if hasattr(ed,'iloc') and len(ed)>0:
                        earnings_date = str(ed.iloc[0].date()) if hasattr(ed.iloc[0],'date') else str(ed.iloc[0])
        except Exception:
            pass

        quality="PASS"; quality_flags=[]
        if pe and pe<0: quality_flags.append("Neg P/E")
        if pe and pe>100: quality_flags.append("P/E>100")
        if rev_growth and rev_growth<-0.1: quality_flags.append("Rev-10%+")
        if debt_equity and debt_equity>300: quality_flags.append("High debt")
        if len(quality_flags)>=2: quality="FAIL"
        elif len(quality_flags)==1: quality="WARN"

        earnings_soon = False
        if earnings_date:
            try:
                ed_dt = datetime.strptime(earnings_date,"%Y-%m-%d")
                if 0<=(ed_dt-datetime.now()).days<=5: earnings_soon=True
            except Exception:
                pass

        return {
            "pe":round(pe,1) if pe else None,
            "rev_growth":round(rev_growth*100,1) if rev_growth else None,
            "profit_margin":round(profit_margin*100,1) if profit_margin else None,
            "debt_equity":round(debt_equity,0) if debt_equity else None,
            "mkt_cap":mkt_cap,
            "earnings_date":earnings_date,
            "earnings_soon":earnings_soon,
            "quality":quality,
            "quality_flags":quality_flags,
            "short_ratio":round(float(short_ratio),1) if short_ratio else None,
            "short_float_pct":round(float(short_float)*100,1) if short_float else None,
        }
    except Exception:
        return {"pe":None,"rev_growth":None,"profit_margin":None,"debt_equity":None,"mkt_cap":None,
                "earnings_date":None,"earnings_soon":False,"quality":"N/A","quality_flags":[],
                "short_ratio":None,"short_float_pct":None}

# ============================================================
# SCORING v3.0 — FLOW-FIRST
# ============================================================
def score_asset(closes, highs, lows, volumes, bench, finra_data=None, rs_pct=50.0):
    """
    100-point scoring breakdown:
      RS    20pts  relative strength vs SPY + cross-sectional percentile rank
      Flow  30pts  CMF + MFI + OBV divergence + Up/Down volume  ← THE REAL MONEY FLOW
      MA    15pts  trend structure (SMA stack)
      PB    10pts  entry timing / pullback quality
      Mom   15pts  RSI zone + recent momentum
      Short 10pts  FINRA off-exchange short pressure + covering signal
    """
    score = 0

    # --- RS (20 pts) ---
    ret1m = calc_return(closes,21); ret3m = calc_return(closes,63)
    sp1 = bench.get("1M",0) or 0; sp3 = bench.get("3M",0) or 0
    rs1 = (ret1m-sp1) if ret1m else 0; rs3 = (ret3m-sp3) if ret3m else 0
    rs_raw    = max(0, min(15, 7.5+(rs1*.4+rs3*.6)*1.2))
    rs_pct_pt = max(0, min(5, (rs_pct/100)*5)) if rs_pct is not None else 2.5
    rs_pts    = round(rs_raw+rs_pct_pt, 1)
    score += rs_pts

    # --- Flow (30 pts) — THE INSTITUTIONAL FINGERPRINT ---
    cmf     = calc_cmf(closes,highs,lows,volumes,20)
    mfi     = calc_mfi(closes,highs,lows,volumes,14)
    obv_sl  = calc_obv_slope(closes,volumes,20)
    obv_div = calc_obv_divergence(closes,volumes,20)
    udv     = calc_updown_vol(closes,volumes,20)

    flow_pts = 15  # neutral base

    if cmf is not None:
        if   cmf > 0.15: flow_pts += 5   # strong accumulation
        elif cmf > 0.05: flow_pts += 3   # mild accumulation
        elif cmf > 0:    flow_pts += 1
        elif cmf < -0.15:flow_pts -= 5   # strong distribution
        elif cmf < -0.05:flow_pts -= 3   # mild distribution
        else:            flow_pts -= 1

    if mfi is not None:
        if   55 <= mfi <= 75: flow_pts += 4   # bullish sweet spot
        elif 45 <= mfi < 55:  flow_pts += 2   # leaning bullish
        elif mfi > 80:        flow_pts += 1   # overbought (slight caution)
        elif mfi < 30:        flow_pts -= 3   # bearish pressure
        elif mfi < 40:        flow_pts -= 1

    div_pts = {"confirmed_bull":4, "bullish":6, "neutral":0, "confirmed_bear":-3, "bearish":-5}
    flow_pts += div_pts.get(obv_div, 0)

    if udv is not None:
        if   udv > 1.8: flow_pts += 4
        elif udv > 1.3: flow_pts += 2
        elif udv < 0.7: flow_pts -= 3
        elif udv < 0.5: flow_pts -= 4

    flow_pts = max(0, min(30, flow_pts))
    score += flow_pts

    # --- MA (15 pts) ---
    sma20  = calc_sma(closes,20); sma50 = calc_sma(closes,50); sma200 = calc_sma(closes,200)
    price  = float(closes[-1])
    ma_pts = 0
    if sma20  and price>sma20:  ma_pts += 4
    if sma50  and price>sma50:  ma_pts += 4
    if sma200 and price>sma200: ma_pts += 4
    if sma20 and sma50 and sma20>sma50: ma_pts += 2
    if sma50 and sma200 and sma50>sma200: ma_pts += 1
    score += ma_pts

    # --- PB (10 pts) ---
    pb_pts = 5
    if sma20 and sma50:
        d20 = ((price-sma20)/sma20)*100
        if  -1<=d20<=2 and price>sma50: pb_pts = 10
        elif d20>8:                      pb_pts = 3
        elif d20<-5:                     pb_pts = 1
    score += pb_pts

    # --- Mom (15 pts) ---
    rsi    = calc_rsi(closes)
    atr    = calc_atr(highs,lows,closes)
    ret1w  = calc_return(closes,5)
    mom_pts = 7
    if rsi:
        if   50<=rsi<=65: mom_pts += 4
        elif 40<=rsi<50:  mom_pts += 2
        elif rsi>75:      mom_pts -= 2
        elif rsi<30:      mom_pts -= 1
    if ret1w and ret1w>0 and ret1m and ret1m>0: mom_pts += 4
    elif ret1w and ret1w<-3: mom_pts -= 2
    mom_pts = max(0, min(15, mom_pts))
    score += mom_pts

    # --- Short / Dark Pool (10 pts) ---
    short_pts = 5  # neutral
    if finra_data:
        sp_ = finra_data.get("short_pct")
        st_ = finra_data.get("short_pct_trend")
        if sp_ is not None:
            if sp_ < 35:   short_pts += 2   # low short pressure
            elif sp_ > 55: short_pts -= 1   # heavy short pressure
        if st_ is not None:
            if   st_ < -5: short_pts += 2   # shorts covering = bullish
            elif st_ >  5: short_pts -= 2   # shorts accelerating = bearish
    short_pts = max(0, min(10, short_pts))
    score += short_pts

    above_20  = bool(sma20  and price>sma20)
    above_50  = bool(sma50  and price>sma50)
    above_200 = bool(sma200 and price>sma200)

    return {
        "score":round(score), "price":round(price,2),
        "ret1w":ret1w, "ret1m":ret1m, "ret3m":ret3m,
        "rsi":rsi, "atr":sf(atr), "vol_ratio":calc_vol_ratio(volumes),
        "sma20":sma20, "sma50":sma50, "sma200":sma200,
        "above_20":above_20, "above_50":above_50, "above_200":above_200,
        # Flow signals
        "cmf":cmf, "mfi":mfi, "obv_slope":obv_sl, "obv_div":obv_div,
        "updown_vol":udv, "rs_pct":rs_pct,
        "finra_short_pct":   finra_data.get("short_pct")   if finra_data else None,
        "finra_short_trend": finra_data.get("short_pct_trend") if finra_data else None,
        "bd":{"RS":round(rs_pts,1),"Flow":round(flow_pts,1),"MA":round(ma_pts,1),
              "PB":round(pb_pts,1),"Mom":round(mom_pts,1),"Short":round(short_pts,1)},
        "bd_max":{"RS":20,"Flow":30,"MA":15,"PB":10,"Mom":15,"Short":10},
    }

def get_signal(sc):
    if sc>=75: return ("STRONG BUY ZONE","#22c55e","rgba(34,197,94,0.12)")
    if sc>=60: return ("ACCUMULATE","#4ade80","rgba(74,222,128,0.10)")
    if sc>=45: return ("WATCH","#fbbf24","rgba(251,191,36,0.10)")
    if sc>=30: return ("CAUTION","#f97316","rgba(249,115,22,0.10)")
    return ("AVOID","#ef4444","rgba(239,68,68,0.10)")

# ============================================================
# DATA
# ============================================================
def fetch_data(symbol):
    try:
        df = yf.Ticker(symbol).history(period="1y", interval="1d", auto_adjust=True)
        if df.empty or len(df)<30: return None
        return {"dates":[d.strftime("%Y-%m-%d") for d in df.index],
                "closes":df["Close"].values.astype(float),
                "highs": df["High"].values.astype(float),
                "lows":  df["Low"].values.astype(float),
                "volumes":df["Volume"].values.astype(float)}
    except Exception as e:
        print(f" fetch error {symbol}: {e}")
        return None

def build_chart_data(data, spy):
    c,v,h,l,dates = data["closes"],data["volumes"],data["highs"],data["lows"],data["dates"]
    n = min(120,len(c))
    c_s,v_s,h_s,l_s,d_s = c[-n:],v[-n:],h[-n:],l[-n:],dates[-n:]
    sma20  = calc_sma_series(c,20)[-n:]
    sma50  = calc_sma_series(c,50)[-n:]
    sma200 = calc_sma_series(c,200)[-n:]
    rsi_s  = calc_rsi_series(c)[-n:]
    vc = ["rgba(100,116,139,0.5)" if i==0 else
          "rgba(74,222,128,0.6)" if c_s[i]>=c_s[i-1] else
          "rgba(248,113,113,0.6)" for i in range(len(c_s))]
    va = []
    for i in range(len(v_s)):
        idx = len(v)-n+i
        va.append(sf(np.mean(v[max(0,idx-49):idx+1])) if idx>=49 else None)
    rs = calc_rs_series(c, spy["closes"])[-n:]
    # Rolling CMF for chart
    cmf_s = []
    for i in range(n):
        idx = len(c)-n+i
        cmf_s.append(calc_cmf(c[:idx+1],h[:idx+1],l[:idx+1],v[:idx+1],20) if idx>=20 else None)
    return {"lb":d_s,"pr":[sf(x) for x in c_s],"s20":sma20,"s50":sma50,"s200":sma200,
            "rsi":rsi_s,"vol":[int(x) for x in v_s],"vc":vc,"va":va,"rs":rs,"cmf":cmf_s}

# ============================================================
# EXCEL DASHBOARD (updated for v3.0 flow signals)
# ============================================================
def update_excel(sectors, stocks, bench, macro, regime, scan_dt, path):
    is_new = not os.path.exists(path)
    wb = openpyxl.Workbook() if is_new else openpyxl.load_workbook(path)

    hf   = Font(name="Arial",size=10,bold=True,color="FFFFFF")
    hfill= PatternFill("solid",fgColor="1a1a2e")
    df_  = Font(name="Arial",size=10)
    gf   = Font(name="Arial",size=10,color="22c55e")
    rf   = Font(name="Arial",size=10,color="ef4444")
    yf_  = Font(name="Arial",size=10,color="fbbf24")
    cf_  = Font(name="Arial",size=10,color="22d3ee")
    ct   = Alignment(horizontal="center",vertical="center")
    bdr  = Border(bottom=Side(style="thin",color="333333"),right=Side(style="thin",color="333333"))

    def sigfont(sc):
        if sc is None: return df_
        if sc>=75: return gf
        if sc>=60: return Font(name="Arial",size=10,color="4ade80")
        if sc>=45: return yf_
        if sc>=30: return Font(name="Arial",size=10,color="f97316")
        return rf

    def cmffont(v):
        if v is None: return df_
        if v>0.05: return gf
        if v<-0.05: return rf
        return df_

    def write_hdr(ws,row,headers):
        for col,h in enumerate(headers,1):
            c=ws.cell(row,col,h); c.font=hf; c.fill=hfill; c.alignment=ct; c.border=bdr

    def write_val(ws,row,col,val,font=None):
        c=ws.cell(row,col,val); c.font=font or df_; c.alignment=ct; c.border=bdr

    # ---- Sheet 1: Macro Dashboard ----
    sn="Macro Dashboard"
    ws=wb.active if is_new else (wb[sn] if sn in wb.sheetnames else wb.create_sheet(sn))
    if is_new: ws.title=sn
    r=ws.max_row+2 if ws.max_row and ws.cell(1,1).value else 1
    ws.cell(r,1,f"SCAN: {scan_dt.strftime('%Y-%m-%d %I:%M %p')}").font=Font(name="Arial",size=11,bold=True,color="a5b4fc")
    r+=1
    write_hdr(ws,r,["Regime","VIX","10Y Yield","USD Index","SPY Price","SPY 50MA","SPY 200MA","Posture","SPY 1M%","SPY 3M%"])
    r+=1
    vix=macro.get("VIX",{}).get("current"); tnx=macro.get("TNX",{}).get("current"); dxy=macro.get("DXY",{}).get("current")
    vals=[regime["regime"],vix,tnx,dxy,regime.get("spy_price"),regime.get("spy_sma50"),regime.get("spy_sma200"),regime["posture"],bench.get("1M"),bench.get("3M")]
    for col,v in enumerate(vals,1): write_val(ws,r,col,v)
    ws.cell(r,1).font=Font(name="Arial",size=10,bold=True)
    for i,w in enumerate([14,8,10,10,10,10,10,55,9,9],1): ws.column_dimensions[get_column_letter(i)].width=w

    # ---- Sheet 2: Sector Tracker ----
    sn2="Sector Tracker"
    ws2=wb[sn2] if sn2 in wb.sheetnames else wb.create_sheet(sn2)
    r2=ws2.max_row+2 if ws2.max_row and ws2.cell(1,1).value else 1
    ws2.cell(r2,1,f"SCAN: {scan_dt.strftime('%Y-%m-%d %I:%M %p')} | Regime: {regime['regime']}").font=Font(name="Arial",size=11,bold=True,color="a5b4fc")
    r2+=1
    write_hdr(ws2,r2,["Rank","Symbol","Sector","Score","Signal","Price","1W%","1M%","3M%","RSI","VolR",
                       "CMF","MFI","OBV Div","UD Vol","Breadth 20MA","Breadth 50MA","RS","Flow","MA","PB","Mom","Short"])
    r2+=1
    for i,s in enumerate(sectors):
        if s.get("error"): continue
        sl,_,_=get_signal(s["score"]); bd=s.get("bd",{})
        b20=s.get("breadth_20",0); b50=s.get("breadth_50",0)
        vals=[i+1,s["symbol"],s["name"],s["score"],sl,s["price"],
              s.get("ret1w"),s.get("ret1m"),s.get("ret3m"),s.get("rsi"),s.get("vol_ratio"),
              s.get("cmf"),s.get("mfi"),s.get("obv_div"),s.get("updown_vol"),
              b20,b50,
              bd.get("RS"),bd.get("Flow"),bd.get("MA"),bd.get("PB"),bd.get("Mom"),bd.get("Short")]
        for col,v in enumerate(vals,1): write_val(ws2,r2,col,v)
        ws2.cell(r2,4).font=sigfont(s["score"]); ws2.cell(r2,5).font=sigfont(s["score"])
        ws2.cell(r2,12).font=cmffont(s.get("cmf"))
        for ci in [7,8,9]:
            v=ws2.cell(r2,ci).value
            if v is not None: ws2.cell(r2,ci).font=gf if v>=0 else rf
        r2+=1
    for i,w in enumerate([5,7,18,7,16,9,8,8,8,6,7,7,6,13,7,12,12,6,7,6,6,6,7],1):
        ws2.column_dimensions[get_column_letter(i)].width=w

    # ---- Sheet 3: Stock Tracker ----
    sn3="Stock Tracker"
    ws3=wb[sn3] if sn3 in wb.sheetnames else wb.create_sheet(sn3)
    r3=ws3.max_row+2 if ws3.max_row and ws3.cell(1,1).value else 1
    ws3.cell(r3,1,f"SCAN: {scan_dt.strftime('%Y-%m-%d %I:%M %p')}").font=Font(name="Arial",size=11,bold=True,color="a5b4fc")
    r3+=1
    write_hdr(ws3,r3,["Sector","Rank","Symbol","Score","Signal","Price","1W%","1M%","3M%","RSI","VolR",
                       "CMF","MFI","OBV Div","UD Vol","RS%ile","FINRA Sh%","Sh Trend",
                       "Stop","T1","T2","P/E","RevGr%","Quality","Earnings","EarnSoon"])
    r3+=1
    for sym,nm,clr in SECTOR_ETFS:
        for j,st in enumerate(stocks.get(sym,[])):
            if st.get("error"): continue
            sl,_,_=get_signal(st["score"]); fu=st.get("fund",{})
            vals=[sym,j+1,st["symbol"],st["score"],sl,st["price"],
                  st.get("ret1w"),st.get("ret1m"),st.get("ret3m"),st.get("rsi"),st.get("vol_ratio"),
                  st.get("cmf"),st.get("mfi"),st.get("obv_div"),st.get("updown_vol"),
                  st.get("rs_pct"),st.get("finra_short_pct"),st.get("finra_short_trend"),
                  st.get("stop"),st.get("t1"),st.get("t2"),
                  fu.get("pe"),fu.get("rev_growth"),fu.get("quality"),
                  fu.get("earnings_date"),"YES" if fu.get("earnings_soon") else ""]
            for col,v in enumerate(vals,1): write_val(ws3,r3,col,v)
            ws3.cell(r3,4).font=sigfont(st["score"]); ws3.cell(r3,5).font=sigfont(st["score"])
            ws3.cell(r3,12).font=cmffont(st.get("cmf"))
            ws3.cell(r3,19).font=rf; ws3.cell(r3,20).font=gf; ws3.cell(r3,21).font=gf
            if fu.get("earnings_soon"): ws3.cell(r3,26).font=Font(name="Arial",size=10,bold=True,color="ef4444")
            qf=fu.get("quality","N/A")
            ws3.cell(r3,24).font=gf if qf=="PASS" else yf_ if qf=="WARN" else rf
            r3+=1
    for i,w in enumerate([7,5,8,7,16,9,8,8,8,6,7,7,6,13,7,7,9,9,9,9,9,7,8,8,12,9],1):
        ws3.column_dimensions[get_column_letter(i)].width=w

    # ---- Sheet 4: Score History ----
    sn4="Score History"
    ws4=wb[sn4] if sn4 in wb.sheetnames else wb.create_sheet(sn4)
    if ws4.max_row is None or ws4.max_row<1 or ws4.cell(1,1).value is None:
        hdrs=["Date","Regime"]+[s[0] for s in SECTOR_ETFS]+["SPY 1M%","SPY 3M%","VIX"]
        write_hdr(ws4,1,hdrs); hr4=2
    else: hr4=ws4.max_row+1
    write_val(ws4,hr4,1,scan_dt.strftime("%Y-%m-%d %H:%M"))
    write_val(ws4,hr4,2,regime["regime"])
    sec_lookup={s["symbol"]:s.get("score",0) for s in sectors if not s.get("error")}
    for col,(sym,_,_) in enumerate(SECTOR_ETFS,3):
        sc=sec_lookup.get(sym)
        write_val(ws4,hr4,col,sc,sigfont(sc) if sc else df_)
    write_val(ws4,hr4,len(SECTOR_ETFS)+3,bench.get("1M"))
    write_val(ws4,hr4,len(SECTOR_ETFS)+4,bench.get("3M"))
    write_val(ws4,hr4,len(SECTOR_ETFS)+5,macro.get("VIX",{}).get("current"))
    ws4.column_dimensions["A"].width=16; ws4.column_dimensions["B"].width=14

    # ---- Sheet 5: Top Movers (60+) ----
    sn5="Top Movers"
    ws5=wb[sn5] if sn5 in wb.sheetnames else wb.create_sheet(sn5)
    r5=ws5.max_row+2 if ws5.max_row and ws5.cell(1,1).value else 1
    ws5.cell(r5,1,f"SCAN: {scan_dt.strftime('%Y-%m-%d %I:%M %p')}").font=Font(name="Arial",size=11,bold=True,color="a5b4fc")
    r5+=1
    write_hdr(ws5,r5,["Rank","Symbol","Sector","Score","Signal","Price","1M%","RSI","CMF","MFI","OBV Div","UD Vol","Quality","Earnings","Stop","T1","T2"])
    r5+=1
    snl={s[0]:s[1] for s in SECTOR_ETFS}; tops=[]
    for sym in SECTOR_STOCKS:
        for st in stocks.get(sym,[]):
            if not st.get("error") and st.get("score",0)>=60:
                st["_sec"]=snl.get(sym,sym); tops.append(st)
    tops.sort(key=lambda x:x.get("score",0),reverse=True)
    for j,st in enumerate(tops):
        sl,_,_=get_signal(st["score"]); fu=st.get("fund",{})
        vals=[j+1,st["symbol"],st["_sec"],st["score"],sl,st["price"],st.get("ret1m"),st.get("rsi"),
              st.get("cmf"),st.get("mfi"),st.get("obv_div"),st.get("updown_vol"),
              fu.get("quality"),fu.get("earnings_date"),st.get("stop"),st.get("t1"),st.get("t2")]
        for col,v in enumerate(vals,1): write_val(ws5,r5,col,v)
        ws5.cell(r5,4).font=sigfont(st["score"]); ws5.cell(r5,9).font=cmffont(st.get("cmf"))
        r5+=1
    if not tops: ws5.cell(r5,1,"No stocks 60+ this scan").font=Font(name="Arial",size=10,color="64748b")

    # ---- Sheet 6: Short Candidates (<30) ----
    sn6="Short Candidates"
    ws6=wb[sn6] if sn6 in wb.sheetnames else wb.create_sheet(sn6)
    r6=ws6.max_row+2 if ws6.max_row and ws6.cell(1,1).value else 1
    ws6.cell(r6,1,f"SCAN: {scan_dt.strftime('%Y-%m-%d %I:%M %p')}").font=Font(name="Arial",size=11,bold=True,color="a5b4fc")
    r6+=1
    write_hdr(ws6,r6,["Rank","Symbol","Sector","Score","Signal","Price","1M%","3M%","RSI","CMF","MFI","OBV Div","UD Vol","FINRA Sh%","Quality"])
    r6+=1
    shorts=[]
    for sym in SECTOR_STOCKS:
        for st in stocks.get(sym,[]):
            if not st.get("error") and st.get("score",0)<30:
                st["_sec"]=snl.get(sym,sym); shorts.append(st)
    shorts.sort(key=lambda x:x.get("score",0))
    for j,st in enumerate(shorts):
        sl,_,_=get_signal(st["score"]); fu=st.get("fund",{})
        vals=[j+1,st["symbol"],st["_sec"],st["score"],sl,st["price"],
              st.get("ret1m"),st.get("ret3m"),st.get("rsi"),
              st.get("cmf"),st.get("mfi"),st.get("obv_div"),st.get("updown_vol"),
              st.get("finra_short_pct"),fu.get("quality")]
        for col,v in enumerate(vals,1): write_val(ws6,r6,col,v)
        ws6.cell(r6,4).font=rf; ws6.cell(r6,10).font=cmffont(st.get("cmf"))
        r6+=1

    # ---- Sheet 7: Trade Journal ----
    sn7="Trade Journal"
    if sn7 not in wb.sheetnames:
        ws7=wb.create_sheet(sn7)
        write_hdr(ws7,1,["Date","Symbol","Direction","Entry Price","Position Size","Score at Entry",
                          "Signal","CMF at Entry","MFI at Entry","OBV Div","Stop Loss",
                          "Target 1","Target 2","Exit Date","Exit Price","P/L $","P/L %",
                          "Reason for Entry","Reason for Exit","Lessons Learned"])
        for i,w in enumerate([12,8,9,11,12,11,16,10,10,13,10,10,10,12,11,9,8,25,25,30],1):
            ws7.column_dimensions[get_column_letter(i)].width=w
        for row in range(2,22):
            for col in range(1,21):
                c=ws7.cell(row,col,""); c.font=df_; c.alignment=ct; c.border=bdr

    wb.save(path)
    print(f"  Dashboard saved: {path}")

# ============================================================
# HTML BUILD HELPERS
# ============================================================
def build_bd_html(bd, sig_c):
    mx={"RS":20,"Flow":30,"MA":15,"PB":10,"Mom":15,"Short":10}
    h=""
    for k,v in bd.items():
        m=mx.get(k,20); p=min(100,v/m*100) if m>0 else 0
        h+=f'<div class="bdi"><div class="bdl">{k}</div><div class="bdo"><div class="bdf" style="width:{p:.0f}%;background:{sig_c}"></div></div><div class="bdv">{v}/{m}</div></div>'
    return h

def _cmf_color(v):
    if v is None: return "#64748b"
    if v>0.1:  return "#22c55e"
    if v>0.02: return "#4ade80"
    if v<-0.1: return "#ef4444"
    if v<-0.02:return "#f97316"
    return "#64748b"

def _div_badge(d):
    labels={"confirmed_bull":("CONF BULL","#22c55e"),"bullish":("HIDDEN ACC","#4ade80"),
            "neutral":("NEUTRAL","#64748b"),"confirmed_bear":("CONF BEAR","#f97316"),
            "bearish":("DISTRIB","#ef4444")}
    lbl,clr = labels.get(d,("?","#64748b"))
    return f'<span style="font-size:7px;font-weight:700;padding:1px 4px;border-radius:2px;background:{clr}20;color:{clr}">{lbl}</span>'

def _mfi_color(v):
    if v is None: return "#64748b"
    if 55<=v<=75: return "#22c55e"
    if v>80 or v<30: return "#ef4444"
    if v>=45: return "#fbbf24"
    return "#f97316"

# ============================================================
# HTML INTERPRETATION GUIDE
# ============================================================
GUIDE_HTML = '''
<div class="tt" id="guide-toggle" onclick="toggleGuide()" style="cursor:pointer">
  How to Read This Report <span id="guide-arrow" style="font-size:11px;color:#64748b"> ▼ tap to expand</span>
</div>
<div id="guide-body" style="display:none">

<div class="guide-section" style="border:2px solid rgba(99,102,241,0.35);background:rgba(99,102,241,0.04)">
  <div class="guide-title" style="color:#a5b4fc">THE RUBIK'S CUBE — 6 FACES MUST ALIGN</div>
  <div style="font-size:10px;color:#cbd5e1;margin-bottom:12px;line-height:1.6">
    Every trade has 6 independent dimensions, like the faces of a Rubik's cube. The cube is "solved" only when every face turns green together. A score of 75+ usually means 5-6 faces aligned. If even one face is red — the cube isn't solved. Don't force a trade.
  </div>
  <div class="cube-grid">
    <div class="cube-face" style="border-color:#6366f1">
      <div class="cf-tag" style="background:#6366f1;color:#fff">RS 20</div>
      <div class="cf-name">Relative Strength</div>
      <div class="cf-q">"Is this stock leading the market?"</div>
      <div class="cf-pts">Return vs SPY + universe rank</div>
    </div>
    <div class="cube-face" style="border-color:#22d3ee">
      <div class="cf-tag" style="background:#22d3ee;color:#000">FLOW 30</div>
      <div class="cf-name">Money Flow ★</div>
      <div class="cf-q">"Where is institutional money going?"</div>
      <div class="cf-pts">CMF + MFI + OBV + UD Volume</div>
    </div>
    <div class="cube-face" style="border-color:#a78bfa">
      <div class="cf-tag" style="background:#a78bfa;color:#000">MA 15</div>
      <div class="cf-name">Moving Averages</div>
      <div class="cf-q">"Is the trend structure intact?"</div>
      <div class="cf-pts">Price above SMA 20/50/200</div>
    </div>
    <div class="cube-face" style="border-color:#fbbf24">
      <div class="cf-tag" style="background:#fbbf24;color:#000">PB 10</div>
      <div class="cf-name">Pullback Quality</div>
      <div class="cf-q">"Is this a clean entry?"</div>
      <div class="cf-pts">Distance from 20MA + RSI cool-off</div>
    </div>
    <div class="cube-face" style="border-color:#34d399">
      <div class="cf-tag" style="background:#34d399;color:#000">MOM 15</div>
      <div class="cf-name">Momentum</div>
      <div class="cf-q">"Is RSI in the sweet spot?"</div>
      <div class="cf-pts">RSI zone + recent price slope</div>
    </div>
    <div class="cube-face" style="border-color:#f97316">
      <div class="cf-tag" style="background:#f97316;color:#fff">SHORT 10</div>
      <div class="cf-name">Short Pressure</div>
      <div class="cf-q">"Are shorts trapped or right?"</div>
      <div class="cf-pts">FINRA short % + trend direction</div>
    </div>
  </div>
</div>

<div class="guide-section">
  <div class="guide-title">THE 4-STEP FRAMEWORK</div>
  <div class="guide-steps">
    <div class="guide-step"><div class="gs-num">1</div><div><div class="gs-h">MACRO REGIME</div><div class="gs-d">Sets your position sizing. Don't fight it. The regime tells you how big to swing.</div></div></div>
    <div class="guide-step"><div class="gs-num">2</div><div><div class="gs-h">SECTOR SCORE + BREADTH</div><div class="gs-d">Find the leading sector. Breadth ≥7/10 above 50MA means the whole sector is moving — not just one name.</div></div></div>
    <div class="guide-step"><div class="gs-num">3</div><div><div class="gs-h">STOCK SCORE + RS %ILE</div><div class="gs-d">Best name in best sector. RS Percentile >75 means top quartile vs all 110 stocks.</div></div></div>
    <div class="guide-step"><div class="gs-num">4</div><div><div class="gs-h">FLOW CONFIRMATION</div><div class="gs-d">CMF, MFI, OBV must agree. If price is up but flow is negative — institutions are distributing into your buy. Don't enter.</div></div></div>
  </div>
</div>

<div class="guide-section">
  <div class="guide-title">SIGNAL DEFINITIONS &amp; QUICK REFERENCE</div>
  <div class="guide-grid">

    <div class="guide-card">
      <div class="gc-title" style="color:#22d3ee">CMF — Chaikin Money Flow (20)</div>
      <div class="gc-body">Measures where close lands within the day's range, weighted by volume. Unlike simple up/down day counts — this captures HOW MUCH buying vs selling occurred each session. Range: -1.0 to +1.0.</div>
      <div class="gc-formula">CMF = Σ[((C-L)-(H-C))/(H-L) × V] / ΣV  over 20 days</div>
      <div class="gc-rows">
        <div class="gc-row"><span class="gc-val green">+0.15 to +1.0</span><span class="gc-label">Strong accumulation — institutions loading</span></div>
        <div class="gc-row"><span class="gc-val green">+0.05 to +0.15</span><span class="gc-label">Mild accumulation — watch for continuation</span></div>
        <div class="gc-row"><span class="gc-val muted">-0.05 to +0.05</span><span class="gc-label">Neutral — no directional conviction</span></div>
        <div class="gc-row"><span class="gc-val red">-0.05 to -0.15</span><span class="gc-label">Mild distribution — caution, reduce size</span></div>
        <div class="gc-row"><span class="gc-val red">-0.15 to -1.0</span><span class="gc-label">Strong distribution — institutions exiting</span></div>
      </div>
    </div>

    <div class="guide-card">
      <div class="gc-title" style="color:#a78bfa">MFI — Money Flow Index (14)</div>
      <div class="gc-body">RSI calculated on (Typical Price × Volume) instead of price alone. Catches overbought/oversold conditions that RSI misses when volume tells a different story. Range: 0 to 100.</div>
      <div class="gc-formula">TP = (H+L+C)/3 ; MoneyFlow = TP × V ; MFI = 100 - 100/(1 + posMF/negMF) over 14d</div>
      <div class="gc-rows">
        <div class="gc-row"><span class="gc-val green">55 – 75</span><span class="gc-label">Bullish sweet spot — strong but not overheated</span></div>
        <div class="gc-row"><span class="gc-val muted">45 – 55</span><span class="gc-label">Leaning bullish — accumulation building</span></div>
        <div class="gc-row"><span class="gc-val red">&gt; 80</span><span class="gc-label">Overbought — institutional selling pressure likely</span></div>
        <div class="gc-row"><span class="gc-val red">&lt; 30</span><span class="gc-label">Bearish flow — selling dominates on volume</span></div>
        <div class="gc-row"><span class="gc-val muted">&lt; 20</span><span class="gc-label">Extreme oversold — potential reversal zone</span></div>
      </div>
    </div>

    <div class="guide-card">
      <div class="gc-title" style="color:#34d399">RSI — Relative Strength Index (14)</div>
      <div class="gc-body">Pure price momentum oscillator over 14 days. Measures speed of price changes — without volume context. Use alongside MFI: RSI says "is price moving fast" and MFI says "is volume backing it up". Range: 0 to 100.</div>
      <div class="gc-formula">RSI = 100 - 100/(1 + avg_gain/avg_loss)  over 14 days</div>
      <div class="gc-rows">
        <div class="gc-row"><span class="gc-val green">55 – 70</span><span class="gc-label">Sweet spot — strong trend, not yet overheated</span></div>
        <div class="gc-row"><span class="gc-val muted">45 – 55</span><span class="gc-label">Neutral / consolidating — wait for breakout</span></div>
        <div class="gc-row"><span class="gc-val red">&gt; 75</span><span class="gc-label">Overbought — pullback risk increasing</span></div>
        <div class="gc-row"><span class="gc-val red">&lt; 30</span><span class="gc-label">Oversold — but watch for reversal divergence</span></div>
        <div class="gc-row"><span class="gc-val green">Bull divergence</span><span class="gc-label">Price lower-low, RSI higher-low = bounce coming</span></div>
      </div>
    </div>

    <div class="guide-card">
      <div class="gc-title" style="color:#fbbf24">OBV Divergence — Smart Money Signal</div>
      <div class="gc-body">On-Balance Volume tracks cumulative buy/sell pressure. When OBV and price disagree, OBV is usually right — it reflects where the money actually went, not just where price landed.</div>
      <div class="gc-formula">OBV today = OBV yesterday ± today's volume (+ if up close, - if down)</div>
      <div class="gc-rows">
        <div class="gc-row"><span class="gc-val" style="color:#22c55e">HIDDEN ACC ★</span><span class="gc-label">Price down, OBV up — institutions buying dips quietly. Highest-conviction entry signal.</span></div>
        <div class="gc-row"><span class="gc-val green">CONF BULL</span><span class="gc-label">Price up, OBV up — trend confirmed, safe to add</span></div>
        <div class="gc-row"><span class="gc-val muted">NEUTRAL</span><span class="gc-label">No meaningful divergence — wait for clarity</span></div>
        <div class="gc-row"><span class="gc-val red">DISTRIB ⚠</span><span class="gc-label">Price up, OBV down — institutions selling into your rally. Do NOT enter.</span></div>
        <div class="gc-row"><span class="gc-val red">CONF BEAR</span><span class="gc-label">Price down, OBV down — confirmed weakness</span></div>
      </div>
    </div>

    <div class="guide-card">
      <div class="gc-title" style="color:#5eead4">Up/Down Volume Ratio (20)</div>
      <div class="gc-body">Raw ratio of volume traded on up-close days vs down-close days over 20 sessions. Separates buying from selling pressure without price direction bias.</div>
      <div class="gc-formula">UD = Σ(volume on up days) / Σ(volume on down days) over 20 sessions</div>
      <div class="gc-rows">
        <div class="gc-row"><span class="gc-val green">&gt; 1.8</span><span class="gc-label">Strong buying dominance — institutional accumulation</span></div>
        <div class="gc-row"><span class="gc-val green">1.3 – 1.8</span><span class="gc-label">Moderate buying edge</span></div>
        <div class="gc-row"><span class="gc-val muted">0.8 – 1.3</span><span class="gc-label">Balanced — no clear directional pressure</span></div>
        <div class="gc-row"><span class="gc-val red">0.5 – 0.8</span><span class="gc-label">Selling pressure — institutions distributing</span></div>
        <div class="gc-row"><span class="gc-val red">&lt; 0.5</span><span class="gc-label">Heavy selling — exit or avoid</span></div>
      </div>
    </div>

    <div class="guide-card">
      <div class="gc-title" style="color:#f97316">FINRA Short % (Dark Pool Proxy)</div>
      <div class="gc-body">Daily short sale volume as % of all FINRA-reported off-exchange trades. Baseline is 40-60% (market maker hedging is normal). Deviations signal directional pressure.</div>
      <div class="gc-formula">Short% = (FINRA short volume / FINRA total volume) × 100, sourced daily</div>
      <div class="gc-rows">
        <div class="gc-row"><span class="gc-val green">&lt; 35%</span><span class="gc-label">Low short pressure — longs in control</span></div>
        <div class="gc-row"><span class="gc-val muted">35 – 55%</span><span class="gc-label">Normal range — market maker activity</span></div>
        <div class="gc-row"><span class="gc-val red">&gt; 55%</span><span class="gc-label">Elevated short pressure OR squeeze fuel (stock dependent)</span></div>
        <div class="gc-row"><span class="gc-val green">Trend ↓ sharply</span><span class="gc-label">Shorts covering = bullish catalyst incoming</span></div>
        <div class="gc-row"><span class="gc-val red">Trend ↑ sharply</span><span class="gc-label">Short sellers accelerating = bearish pressure building</span></div>
      </div>
    </div>

    <div class="guide-card">
      <div class="gc-title" style="color:#6366f1">RS Percentile Rank</div>
      <div class="gc-body">Where this stock ranks vs all 110 stocks in the universe by 1-month return. Replaces raw return-vs-SPY with a true ranking signal.</div>
      <div class="gc-formula">RS%ile = percentile rank of (1-month return) across all 110 stocks in universe</div>
      <div class="gc-rows">
        <div class="gc-row"><span class="gc-val green">75 – 100</span><span class="gc-label">Top quartile — price leaders, institution favorites</span></div>
        <div class="gc-row"><span class="gc-val green">50 – 75</span><span class="gc-label">Above average — solid momentum</span></div>
        <div class="gc-row"><span class="gc-val muted">25 – 50</span><span class="gc-label">Below average — avoid or wait for rotation</span></div>
        <div class="gc-row"><span class="gc-val red">&lt; 25</span><span class="gc-label">Laggard — institutional money is leaving</span></div>
      </div>
    </div>

  </div>
</div>

<div class="guide-section">
  <div class="guide-title">CONFLUENCE RULES — WHEN TO ACT</div>
  <div class="guide-grid2">
    <div class="guide-rule green-rule">
      <div class="gr-title">GREEN LIGHT (all 4 must align)</div>
      <div class="gr-items">
        <div>✓ Macro: TRENDING UP or MODERATE</div>
        <div>✓ Sector score ≥ 60, Breadth ≥ 6/10</div>
        <div>✓ Stock score ≥ 65, RS Percentile ≥ 70</div>
        <div>✓ CMF &gt; +0.05  AND  MFI 50-75  AND  OBV not BEARISH</div>
        <div>✓ No earnings within 5 days</div>
      </div>
    </div>
    <div class="guide-rule red-rule">
      <div class="gr-title">RED FLAGS (any one = no entry)</div>
      <div class="gr-items">
        <div>✗ OBV Divergence = BEARISH (distribution trap)</div>
        <div>✗ CMF &lt; -0.1 (institutions actively selling)</div>
        <div>✗ FINRA Short% trend rising sharply &gt;+8</div>
        <div>✗ Earnings within 5 days (binary event)</div>
        <div>✗ Regime = RISK-OFF / BEAR (no new longs)</div>
        <div>✗ Quality = FAIL (fundamental rot)</div>
      </div>
    </div>
    <div class="guide-rule yellow-rule">
      <div class="gr-title">WATCH LIST CRITERIA</div>
      <div class="gr-items">
        <div>→ OBV = HIDDEN ACCUMULATION + CMF turning positive</div>
        <div>→ Score rising on consecutive scans</div>
        <div>→ FINRA Short% declining (shorts covering)</div>
        <div>→ Sector breadth improving week over week</div>
        <div>→ RS Percentile crossing above 50 from below</div>
      </div>
    </div>
  </div>
</div>

<div class="guide-section">
  <div class="guide-title">SCORE BREAKDOWN LEGEND</div>
  <div style="display:flex;gap:12px;flex-wrap:wrap;margin-top:8px">
    <div class="score-pill"><span style="color:#6366f1">RS 0-20</span> Return vs SPY + universe rank</div>
    <div class="score-pill"><span style="color:#22d3ee">Flow 0-30</span> CMF+MFI+OBV+UD Vol — real money direction</div>
    <div class="score-pill"><span style="color:#a78bfa">MA 0-15</span> Price above SMA 20/50/200 stack</div>
    <div class="score-pill"><span style="color:#fbbf24">PB 0-10</span> Pullback quality / entry timing</div>
    <div class="score-pill"><span style="color:#34d399">Mom 0-15</span> RSI zone + recent price momentum</div>
    <div class="score-pill"><span style="color:#f97316">Short 0-10</span> FINRA short pressure signal</div>
  </div>
</div>

</div>
'''

# ============================================================
# HTML REPORT
# ============================================================
def generate_html(sectors, stocks, bench, macro, regime, scan_time, charts):
    chart_json = safe_json(charts)
    bench_html = "".join(
        f'<span><span class="bk">{k}:</span><span class="bv {pc(v)}">{fmt_pct(v)}</span></span>'
        for k,v in bench.items()
    )

    vix=macro.get("VIX",{}).get("current"); tnx=macro.get("TNX",{}).get("current"); dxy=macro.get("DXY",{}).get("current")
    macro_html=f'''<div class="cd" style="border-color:{regime['color']}40">
<div class="r"><div><div class="sym" style="color:{regime['color']}">{regime['regime']}</div><div class="nm">{regime['desc']}</div></div>
<div style="text-align:right"><div class="sv" style="color:{regime['color']}">VIX: {fmt(vix,1)}</div></div></div>
<div class="mid">
<div class="st"><div class="sl">10Y Yield</div><div class="sv muted">{fmt(tnx,2)}%</div></div>
<div class="st"><div class="sl">USD Index</div><div class="sv muted">{fmt(dxy,1)}</div></div>
<div class="st"><div class="sl">SPY</div><div class="sv">${fmt(regime.get('spy_price'))}</div></div>
<div class="st"><div class="sl">SPY 50MA</div><div class="sv muted">${fmt(regime.get('spy_sma50'))}</div></div>
<div class="st"><div class="sl">SPY 200MA</div><div class="sv muted">${fmt(regime.get('spy_sma200'))}</div></div>
</div><div style="margin-top:8px;font-size:10px;color:{regime['color']};font-weight:600">POSTURE: {regime['posture']}</div></div>'''

    # Sector cards
    cards=""
    for i,s in enumerate(sectors):
        if s.get("error"):
            cards+=f'<div class="cd"><span class="red">{s["symbol"]}: Failed</span></div>'; continue
        sl,sc2,sb=get_signal(s["score"]); bd_html=build_bd_html(s.get("bd",{}),sc2)
        b20=s.get("breadth_20",0) if isinstance(s.get("breadth_20"),int) else 0
        b50=s.get("breadth_50",0) if isinstance(s.get("breadth_50"),int) else 0
        bclr="#4ade80" if b50>=7 else "#fbbf24" if b50>=4 else "#f87171"
        cmf_c=_cmf_color(s.get("cmf")); div_b=_div_badge(s.get("obv_div","neutral"))
        mfi_c=_mfi_color(s.get("mfi"))
        cards+=f'''<div class="cd ck" onclick="T('{s['symbol']}')">
<div class="r"><div class="ri2"><div class="bar" style="background:{s['color']}"></div><div><div class="sym">{s['symbol']} <span class="rk">#{i+1}</span></div><div class="nm">{s['name']}</div></div></div>
<div style="text-align:right"><div class="sc" style="color:{sc2}">{s['score']}</div><div class="sg" style="background:{sb};color:{sc2}">{sl}</div></div></div>
<div class="mid">
<div class="st"><div class="sl">Price</div><div class="sv">${fmt(s['price'])}</div></div>
<div class="st"><div class="sl">1W</div><div class="sv {pc(s['ret1w'])}">{fmt_pct(s['ret1w'])}</div></div>
<div class="st"><div class="sl">1M</div><div class="sv {pc(s['ret1m'])}">{fmt_pct(s['ret1m'])}</div></div>
<div class="st"><div class="sl">3M</div><div class="sv {pc(s['ret3m'])}">{fmt_pct(s['ret3m'])}</div></div>
<div class="st"><div class="sl">RSI</div><div class="sv muted">{fmt(s['rsi'],0)}</div></div>
<div class="st"><div class="sl">Breadth</div><div class="sv" style="color:{bclr}">{b50}/10 >50MA</div></div>
</div>
<div class="flow-row">
<div class="fitem"><div class="fl">CMF</div><div class="fv" style="color:{cmf_c}">{fmt(s.get('cmf'),3)}</div></div>
<div class="fitem"><div class="fl">MFI</div><div class="fv" style="color:{mfi_c}">{fmt(s.get('mfi'),0)}</div></div>
<div class="fitem"><div class="fl">OBV</div><div class="fv">{div_b}</div></div>
<div class="fitem"><div class="fl">UD Vol</div><div class="fv {'green' if (s.get('updown_vol') or 0)>1.3 else 'red' if (s.get('updown_vol') or 1)<0.8 else 'muted'}">{fmt(s.get('updown_vol'),2)}x</div></div>
</div>
<div class="bd">{bd_html}</div>
<div class="sbo"><div class="sbi" style="width:{s['score']}%;background:{sc2}"></div></div>
<div class="hint">tap for charts</div></div>
<div class="cp" id="c-{s['symbol']}">
  <canvas id="p-{s['symbol']}" height="160"></canvas>
  <canvas id="r-{s['symbol']}" height="100"></canvas>
  <canvas id="mf-{s['symbol']}" height="100"></canvas>
  <canvas id="v-{s['symbol']}" height="100"></canvas>
  <canvas id="s-{s['symbol']}" height="100"></canvas>
</div>'''

    # Stock cards
    stk_html=""
    for sym,name,color in SECTOR_ETFS:
        sl2=stocks.get(sym,[])
        if not sl2: continue
        ss=next((x for x in sectors if x.get("symbol")==sym and not x.get("error")),None)
        ssc=ss["score"] if ss else 0; ssl,ssc2,ssb=get_signal(ssc)
        stk_html+=f'<div class="dv"></div><div class="dh"><div class="dhb" style="background:{color}"></div><div><div class="dhs">{sym} - {name}</div><div class="dhn">Top 10 | Sector Score: {ssc}/100</div></div><div style="margin-left:auto"><span class="sg" style="background:{ssb};color:{ssc2}">{ssl}</span></div></div>'
        for j,st in enumerate(sl2):
            if st.get("error"):
                stk_html+=f'<div class="cd"><span class="red">{st["symbol"]}: Failed</span></div>'; continue
            tl,tc,tb=get_signal(st["score"]); bd2=build_bd_html(st.get("bd",{}),tc)
            fu=st.get("fund",{})
            cmf_c=_cmf_color(st.get("cmf")); div_b=_div_badge(st.get("obv_div","neutral"))
            mfi_c=_mfi_color(st.get("mfi"))

            earn_html=""
            if fu.get("earnings_soon"):
                earn_html=f'<div style="margin-top:6px;padding:4px 8px;background:rgba(239,68,68,0.1);border:1px solid rgba(239,68,68,0.3);border-radius:4px;font-size:9px;color:#f87171;font-weight:700">EARNINGS WITHIN 5 DAYS ({fu.get("earnings_date","?")}). Binary risk — no standard momentum sizing.</div>'
            elif fu.get("earnings_date"):
                earn_html=f'<div style="font-size:8px;color:#64748b;margin-top:4px">Earnings: {fu["earnings_date"]}</div>'

            q=fu.get("quality","N/A"); qclr="#22c55e" if q=="PASS" else "#fbbf24" if q=="WARN" else "#ef4444"
            qual_html=f'<span style="font-size:8px;font-weight:700;padding:1px 5px;border-radius:2px;background:{qclr}20;color:{qclr}">{q}</span>'
            parts=[]
            if fu.get("pe"): parts.append(f'P/E:{fu["pe"]}')
            if fu.get("rev_growth") is not None: parts.append(f'Rev:{"+" if fu["rev_growth"]>=0 else ""}{fu["rev_growth"]}%')
            if fu.get("short_float_pct"): parts.append(f'Short:{fu["short_float_pct"]}%')
            fund_line=f'<div style="font-size:8px;color:#64748b;margin-top:4px">{" | ".join(parts)} | {qual_html}</div>' if parts else ""

            rs_pct=st.get("rs_pct"); rs_clr="#22c55e" if rs_pct and rs_pct>=75 else "#fbbf24" if rs_pct and rs_pct>=50 else "#f87171"
            finra_s=st.get("finra_short_pct"); finra_html=""
            if finra_s is not None:
                ft=st.get("finra_short_trend",0) or 0
                ft_s=f"{'↑' if ft>0 else '↓'}{abs(ft):.1f}%" if ft else ""
                fc="#22c55e" if finra_s<35 else "#ef4444" if finra_s>55 else "#64748b"
                finra_html=f'<div class="fitem"><div class="fl">FINRA Sh%</div><div class="fv" style="color:{fc}">{fmt(finra_s,1)}% {ft_s}</div></div>'

            stk_html+=f'''<div class="cd ck" onclick="T('{st['symbol']}')">
<div class="r"><div><span class="sym">{st['symbol']}</span><span class="rk" style="margin-left:6px">#{j+1}</span></div>
<div style="text-align:right"><div class="sv">${fmt(st['price'])}</div><div class="sg" style="background:{tb};color:{tc}">{tl}</div></div></div>
<div style="display:flex;align-items:center;gap:8px;margin-top:8px">
  <span style="font-weight:700;font-size:16px;color:{tc}">{st['score']}</span>
  <div style="flex:1"><div class="sbo"><div class="sbi" style="width:{st['score']}%;background:{tc}"></div></div></div>
  <span style="font-size:9px;color:{rs_clr}">RS%ile: {fmt(rs_pct,0)}</span>
</div>
<div class="grd">
<div><div class="gl">1W</div><div class="gv {pc(st['ret1w'])}">{fmt_pct(st['ret1w'])}</div></div>
<div><div class="gl">1M</div><div class="gv {pc(st['ret1m'])}">{fmt_pct(st['ret1m'])}</div></div>
<div><div class="gl">3M</div><div class="gv {pc(st['ret3m'])}">{fmt_pct(st['ret3m'])}</div></div>
<div><div class="gl">RSI</div><div class="gv muted">{fmt(st['rsi'],0)}</div></div>
</div>
<div class="flow-row">
<div class="fitem"><div class="fl">CMF</div><div class="fv" style="color:{cmf_c}">{fmt(st.get('cmf'),3)}</div></div>
<div class="fitem"><div class="fl">MFI</div><div class="fv" style="color:{mfi_c}">{fmt(st.get('mfi'),0)}</div></div>
<div class="fitem"><div class="fl">OBV</div><div class="fv">{div_b}</div></div>
<div class="fitem"><div class="fl">UD Vol</div><div class="fv {'green' if (st.get('updown_vol') or 0)>1.3 else 'red' if (st.get('updown_vol') or 1)<0.8 else 'muted'}">{fmt(st.get('updown_vol'),2)}x</div></div>
{finra_html}
</div>
<div class="bd">{bd2}</div>{fund_line}{earn_html}
<div class="rr">
<div class="rit"><div class="rl">Stop</div><div class="rv red">${fmt(st.get('stop'))}</div></div>
<div class="rit"><div class="rl">T1</div><div class="rv green">${fmt(st.get('t1'))}</div></div>
<div class="rit"><div class="rl">T2</div><div class="rv green">${fmt(st.get('t2'))}</div></div>
</div><div class="hint">tap for charts</div></div>
<div class="cp" id="c-{st['symbol']}">
  <canvas id="p-{st['symbol']}" height="160"></canvas>
  <canvas id="r-{st['symbol']}" height="100"></canvas>
  <canvas id="mf-{st['symbol']}" height="100"></canvas>
  <canvas id="v-{st['symbol']}" height="100"></canvas>
  <canvas id="s-{st['symbol']}" height="100"></canvas>
</div>'''

    # Short candidates
    snl={s[0]:s[1] for s in SECTOR_ETFS}
    shorts=[st for sym in SECTOR_STOCKS for st in stocks.get(sym,[]) if not st.get("error") and st.get("score",0)<30]
    shorts.sort(key=lambda x:x.get("score",0))
    short_html=""
    if shorts:
        for j,st in enumerate(shorts[:10]):
            sl3,sc3,sb3=get_signal(st["score"])
            cmf_c=_cmf_color(st.get("cmf")); div_b=_div_badge(st.get("obv_div","neutral"))
            short_html+=f'''<div class="cd" style="border-left:3px solid #ef4444">
<div class="r"><div><span class="sym">{st["symbol"]}</span><span class="rk" style="margin-left:6px">#{j+1}</span></div>
<div style="text-align:right"><span style="font-weight:700;color:#ef4444">{st["score"]}</span> <span class="sg" style="background:{sb3};color:{sc3}">{sl3}</span></div></div>
<div class="flow-row" style="margin-top:8px">
<div class="fitem"><div class="fl">CMF</div><div class="fv" style="color:{cmf_c}">{fmt(st.get("cmf"),3)}</div></div>
<div class="fitem"><div class="fl">OBV</div><div class="fv">{div_b}</div></div>
<div class="fitem"><div class="fl">1M</div><div class="fv {pc(st.get("ret1m"))}">{fmt_pct(st.get("ret1m"))}</div></div>
<div class="fitem"><div class="fl">RSI</div><div class="fv muted">{fmt(st.get("rsi"),0)}</div></div>
</div></div>'''
    else:
        short_html='<div class="cd"><span class="muted">No stocks below 30 this scan</span></div>'

    js_code = '''
var D=''' + chart_json + ''';var R={};
Chart.defaults.color='#64748b';Chart.defaults.borderColor='rgba(255,255,255,0.04)';
Chart.defaults.font.family="'JetBrains Mono',monospace";Chart.defaults.font.size=9;
function toggleGuide(){var b=document.getElementById('guide-body'),a=document.getElementById('guide-arrow');if(b.style.display==='none'){b.style.display='block';a.textContent=' ▲ tap to collapse'}else{b.style.display='none';a.textContent=' ▼ tap to expand'}}
function T(sym){var p=document.getElementById('c-'+sym);if(!p)return;if(p.style.display==='none'||p.style.display===''){p.style.display='block';if(!R[sym]){draw(sym);R[sym]=true}}else{p.style.display='none'}}
function draw(sym){var d=D[sym];if(!d)return;var lb=d.lb.map(function(l){return l.slice(5)});
var opts=function(t){return{responsive:true,plugins:{title:{display:true,text:t,color:'#94a3b8',font:{size:10,weight:600}},legend:{labels:{boxWidth:12,padding:8,font:{size:8}}}},scales:{x:{ticks:{maxTicksLimit:8,font:{size:7}}},y:{position:'right',ticks:{font:{size:8}}}}}};
new Chart(document.getElementById('p-'+sym),{type:'line',data:{labels:lb,datasets:[
{label:'Price',data:d.pr,borderColor:'#e2e8f0',borderWidth:1.5,pointRadius:0,tension:0.1},
{label:'20 SMA',data:d.s20,borderColor:'#6366f1',borderWidth:1,pointRadius:0,borderDash:[3,2]},
{label:'50 SMA',data:d.s50,borderColor:'#f97316',borderWidth:1,pointRadius:0,borderDash:[3,2]},
{label:'200 SMA',data:d.s200,borderColor:'#ef4444',borderWidth:1,pointRadius:0,borderDash:[5,3]}
]},options:opts(sym+' - Price + Moving Averages')});
var ro=opts('RSI (14)');ro.scales.y={min:0,max:100,position:'right',ticks:{stepSize:25,font:{size:8}}};ro.plugins.legend={display:false};
new Chart(document.getElementById('r-'+sym),{type:'line',data:{labels:lb,datasets:[{label:'RSI',data:d.rsi,borderColor:'#a78bfa',borderWidth:1.5,pointRadius:0}]},options:ro,plugins:[{id:'rb',beforeDraw:function(ch){var ctx=ch.ctx,yA=ch.scales.y,xA=ch.scales.x;ctx.fillStyle='rgba(239,68,68,0.06)';ctx.fillRect(xA.left,yA.getPixelForValue(100),xA.width,yA.getPixelForValue(70)-yA.getPixelForValue(100));ctx.fillStyle='rgba(74,222,128,0.06)';ctx.fillRect(xA.left,yA.getPixelForValue(65),xA.width,yA.getPixelForValue(50)-yA.getPixelForValue(65));ctx.strokeStyle='rgba(239,68,68,0.3)';ctx.lineWidth=0.5;ctx.beginPath();ctx.moveTo(xA.left,yA.getPixelForValue(70));ctx.lineTo(xA.right,yA.getPixelForValue(70));ctx.stroke();ctx.strokeStyle='rgba(251,191,36,0.3)';ctx.beginPath();ctx.moveTo(xA.left,yA.getPixelForValue(30));ctx.lineTo(xA.right,yA.getPixelForValue(30));ctx.stroke()}}]});
var cmfClrs=d.cmf.map(function(v){return v===null?'transparent':v>0?'rgba(34,197,94,0.7)':'rgba(239,68,68,0.7)'});
var mfo=opts('CMF — Chaikin Money Flow (20)');mfo.scales.y={position:'right',ticks:{font:{size:8}}};mfo.plugins.legend={display:false};
new Chart(document.getElementById('mf-'+sym),{type:'bar',data:{labels:lb,datasets:[{label:'CMF',data:d.cmf,backgroundColor:cmfClrs,borderWidth:0}]},options:mfo,plugins:[{id:'zl',beforeDraw:function(ch){var ctx=ch.ctx,yA=ch.scales.y,xA=ch.scales.x,y=yA.getPixelForValue(0);if(y){ctx.strokeStyle='rgba(255,255,255,0.2)';ctx.lineWidth=0.8;ctx.setLineDash([]);ctx.beginPath();ctx.moveTo(xA.left,y);ctx.lineTo(xA.right,y);ctx.stroke()}}}]});
var vo=opts('Volume');vo.scales.y.ticks={font:{size:8},callback:function(v){if(v>=1e9)return(v/1e9).toFixed(1)+'B';if(v>=1e6)return(v/1e6).toFixed(0)+'M';if(v>=1e3)return(v/1e3).toFixed(0)+'K';return v}};
new Chart(document.getElementById('v-'+sym),{type:'bar',data:{labels:lb,datasets:[{label:'Volume',data:d.vol,backgroundColor:d.vc,borderWidth:0},{label:'50d Avg',data:d.va,type:'line',borderColor:'#f97316',borderWidth:1.5,pointRadius:0,borderDash:[3,2]}]},options:vo});
var so=opts('RS vs SPY (normalized)');so.plugins.legend={display:false};
new Chart(document.getElementById('s-'+sym),{type:'line',data:{labels:lb,datasets:[{label:'RS',data:d.rs,borderColor:'#22d3ee',borderWidth:1.5,pointRadius:0,fill:{target:'origin',above:'rgba(34,211,238,0.05)',below:'rgba(239,68,68,0.05)'}}]},options:so,plugins:[{id:'bl',beforeDraw:function(ch){var ctx=ch.ctx,yA=ch.scales.y,xA=ch.scales.x,y=yA.getPixelForValue(100);if(y){ctx.strokeStyle='rgba(255,255,255,0.15)';ctx.lineWidth=0.5;ctx.setLineDash([4,4]);ctx.beginPath();ctx.moveTo(xA.left,y);ctx.lineTo(xA.right,y);ctx.stroke();ctx.setLineDash([])}}}]});}
'''

    return f'''<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Money Flow v3 - {scan_time}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<style>@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&display=swap');
*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:'JetBrains Mono',monospace;background:#080c14;color:#e2e8f0;padding:16px}}
.hdr{{border-bottom:1px solid rgba(255,255,255,0.06);padding-bottom:16px;margin-bottom:16px}}.hdr h1{{font-size:15px;font-weight:700;letter-spacing:.08em;color:#f8fafc;text-transform:uppercase}}.hdr .sub{{font-size:10px;color:#64748b;margin-top:4px}}
.cd{{background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.05);border-radius:6px;padding:12px;margin-bottom:2px;font-size:11px}}.ck{{cursor:pointer;transition:background .15s}}.ck:hover{{background:rgba(99,102,241,0.05)}}
.bnch{{display:flex;gap:16px;align-items:center;flex-wrap:wrap;margin-bottom:8px}}.bk{{font-size:9px;color:#64748b;margin-right:4px}}.bv{{font-size:12px;font-weight:600}}
.r{{display:flex;justify-content:space-between;align-items:flex-start}}.ri2{{display:flex;align-items:center;gap:8px}}.bar{{width:3px;height:28px;border-radius:2px}}.sym{{font-weight:700;font-size:13px}}.nm{{font-size:10px;color:#64748b}}.sc{{font-weight:700;font-size:18px}}.rk{{font-size:10px;color:#64748b;font-weight:400}}
.sg{{font-size:8px;font-weight:700;letter-spacing:.06em;padding:2px 6px;border-radius:3px;white-space:nowrap;display:inline-block;margin-top:4px}}
.mid{{display:flex;gap:12px;margin-top:10px;flex-wrap:wrap}}.st .sl{{font-size:8px;color:#64748b;text-transform:uppercase;letter-spacing:.06em}}.st .sv{{font-size:12px;font-weight:600;margin-top:2px}}
.flow-row{{display:flex;gap:10px;margin-top:10px;flex-wrap:wrap;padding:8px;background:rgba(34,211,238,0.03);border:1px solid rgba(34,211,238,0.08);border-radius:4px}}
.fitem{{display:flex;flex-direction:column;min-width:50px}}.fl{{font-size:7px;color:#64748b;text-transform:uppercase;letter-spacing:.06em}}.fv{{font-size:11px;font-weight:600;margin-top:2px}}
.sbo{{width:100%;height:4px;border-radius:2px;background:rgba(255,255,255,0.06);margin-top:8px}}.sbi{{height:100%;border-radius:2px}}.hint{{font-size:8px;color:#475569;text-align:center;margin-top:6px}}
.cp{{background:rgba(255,255,255,0.01);border:1px solid rgba(255,255,255,0.04);border-top:none;border-radius:0 0 6px 6px;padding:12px;margin-bottom:10px;display:none}}.cp canvas{{margin-bottom:12px}}
.bd{{display:flex;gap:6px;flex-wrap:wrap;margin-top:8px}}.bdi{{flex:1;min-width:55px}}.bdl{{font-size:7px;color:#64748b;text-transform:uppercase;letter-spacing:.04em}}.bdo{{width:100%;height:3px;border-radius:2px;background:rgba(255,255,255,0.06);margin-top:2px}}.bdf{{height:100%;border-radius:2px}}.bdv{{font-size:8px;font-weight:600;color:#94a3b8;margin-top:1px}}
.grd{{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-top:10px}}.gl{{font-size:8px;color:#64748b;text-transform:uppercase;letter-spacing:.06em}}.gv{{font-size:11px;font-weight:600;margin-top:2px}}
.rr{{display:flex;gap:10px;margin-top:10px;padding-top:8px;border-top:1px solid rgba(255,255,255,0.04)}}.rit .rl{{font-size:8px;color:#64748b;text-transform:uppercase}}.rit .rv{{font-size:11px;font-weight:600;margin-top:2px}}
.dv{{height:1px;background:rgba(255,255,255,0.06);margin:24px 0}}.dh{{display:flex;align-items:center;gap:10px;margin-bottom:16px;flex-wrap:wrap}}.dhb{{width:4px;height:28px;border-radius:2px}}.dhs{{font-weight:700;font-size:14px}}.dhn{{font-size:10px;color:#64748b}}
.tt{{font-size:13px;font-weight:700;letter-spacing:.06em;color:#f8fafc;text-transform:uppercase;margin:24px 0 12px}}
.ft{{margin-top:32px;padding:12px 0;border-top:1px solid rgba(255,255,255,0.04);font-size:9px;color:#334155}}
.green{{color:#4ade80}}.red{{color:#f87171}}.muted{{color:#94a3b8}}
/* GUIDE STYLES */
.guide-section{{background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.06);border-radius:6px;padding:16px;margin-bottom:10px}}
.guide-title{{font-size:10px;font-weight:700;color:#94a3b8;letter-spacing:.1em;text-transform:uppercase;margin-bottom:12px}}
.guide-steps{{display:flex;flex-direction:column;gap:10px}}
.guide-step{{display:flex;gap:12px;align-items:flex-start}}.gs-num{{width:24px;height:24px;border-radius:50%;background:#6366f1;color:#fff;font-size:11px;font-weight:700;display:flex;align-items:center;justify-content:center;flex-shrink:0}}.gs-h{{font-size:11px;font-weight:700;color:#e2e8f0;margin-bottom:3px}}.gs-d{{font-size:10px;color:#94a3b8}}
.guide-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:10px;margin-top:8px}}
.guide-card{{background:rgba(255,255,255,0.02);border:1px solid rgba(255,255,255,0.06);border-radius:6px;padding:12px}}
.gc-title{{font-size:11px;font-weight:700;margin-bottom:8px}}.gc-body{{font-size:9px;color:#94a3b8;margin-bottom:10px;line-height:1.5}}
.gc-rows{{display:flex;flex-direction:column;gap:5px}}.gc-row{{display:flex;gap:10px;align-items:baseline}}.gc-val{{font-size:9px;font-weight:700;min-width:90px;flex-shrink:0}}.gc-label{{font-size:9px;color:#64748b}}
.guide-grid2{{display:grid;grid-template-columns:repeat(auto-fill,minmax(250px,1fr));gap:10px;margin-top:8px}}
.guide-rule{{border-radius:6px;padding:12px}}.green-rule{{background:rgba(34,197,94,0.05);border:1px solid rgba(34,197,94,0.2)}}.red-rule{{background:rgba(239,68,68,0.05);border:1px solid rgba(239,68,68,0.2)}}.yellow-rule{{background:rgba(251,191,36,0.05);border:1px solid rgba(251,191,36,0.2)}}
.gr-title{{font-size:10px;font-weight:700;color:#94a3b8;letter-spacing:.06em;margin-bottom:8px}}.gr-items{{display:flex;flex-direction:column;gap:5px;font-size:9px;color:#94a3b8}}
.score-pill{{font-size:9px;color:#64748b;background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.06);border-radius:4px;padding:4px 8px}}
.cube-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:8px;margin-top:8px}}
.cube-face{{border:1px solid;border-radius:6px;padding:10px;background:rgba(255,255,255,0.02);position:relative}}
.cf-tag{{position:absolute;top:8px;right:8px;font-size:8px;font-weight:700;color:#000;padding:2px 6px;border-radius:3px}}
.cf-name{{font-size:11px;font-weight:700;color:#e2e8f0;margin-bottom:6px;padding-right:50px}}
.cf-q{{font-size:9px;color:#94a3b8;font-style:italic;margin-bottom:6px;line-height:1.4}}
.cf-pts{{font-size:8px;color:#64748b;text-transform:uppercase;letter-spacing:.05em}}
.gc-formula{{font-size:8px;color:#cbd5e1;font-family:monospace;background:rgba(0,0,0,0.25);padding:5px 8px;border-radius:3px;margin-bottom:8px;border-left:2px solid rgba(99,102,241,0.5)}}
</style></head><body>
<div class="hdr"><h1>Money Flow Screener v3.0 — Institutional Flow Edition</h1><div class="sub">{scan_time} | Yahoo Finance + FINRA Short Data | Not financial advice</div></div>
<div class="tt">Macro Regime</div>{macro_html}
<div class="cd bnch"><span style="font-size:9px;font-weight:700;color:#64748b;letter-spacing:.1em">SPY</span>{bench_html}</div>
{GUIDE_HTML}
<div class="tt">Sector Rankings</div>{cards}
<div class="tt">Stock Drill-Down</div>{stk_html}
<div class="tt">Short Candidates (Score &lt; 30)</div>{short_html}
<div class="ft">v3.0 | Signals: CMF, MFI, OBV Divergence, Up/Down Volume, FINRA Short%, Cross-Sectional RS Rank, SQLite History | Not financial advice.</div>
<script>{js_code}</script></body></html>'''

# ============================================================
# MAIN
# ============================================================
def main():
    scan_dt  = datetime.now()
    ts       = scan_dt.strftime("%Y%m%d_%H%M")
    date_str = scan_dt.strftime("%Y-%m-%d")
    print(f"\n  MONEY FLOW SCREENER v3.0 | {scan_dt.strftime('%B %d, %Y %I:%M %p')}\n")

    os.makedirs("reports", exist_ok=True)
    db = ScoreDB()

    # [1] Macro
    print("  [1/6] Macro environment...")
    macro = fetch_macro()

    # [2] SPY benchmark
    print("\n  [2/6] SPY benchmark...")
    spy = fetch_data(BENCHMARK)
    if not spy: print("  SPY FAILED"); sys.exit(1)
    bench  = {"1W":calc_return(spy["closes"],5),"1M":calc_return(spy["closes"],21),"3M":calc_return(spy["closes"],63)}
    regime = classify_regime(macro, spy)
    print(f"    SPY: ${round(float(spy['closes'][-1]),2)} | Regime: {regime['regime']}")
    print(f"    Posture: {regime['posture']}\n")

    charts = {}

    # [3] Pre-fetch all stock data for cross-sectional RS rank
    print("  [3/6] Pre-fetching stock universe for cross-sectional RS rank...")
    all_symbols = [tk for _sym,_nm,_clr in SECTOR_ETFS for tk in SECTOR_STOCKS.get(_sym,[])]
    all_stock_data  = {}
    all_ret1m       = {}
    for tk in all_symbols:
        d = fetch_data(tk)
        if d:
            all_stock_data[tk] = d
            all_ret1m[tk]      = calc_return(d["closes"],21)
        time.sleep(0.08)
    all_returns_list = list(all_ret1m.values())
    rs_percentiles   = {tk: calc_rs_percentile(ret, all_returns_list) for tk,ret in all_ret1m.items()}
    print(f"    Loaded {len(all_stock_data)}/{len(all_symbols)} stocks")

    # [4] FINRA dark pool / short pressure data
    print("\n  [4/6] FINRA short sale volume (off-exchange)...")
    sector_syms  = [s[0] for s in SECTOR_ETFS]
    all_for_finra = sector_syms + all_symbols
    finra = fetch_finra_short_volume(all_for_finra, lookback_days=5)

    # [5] Sector ETFs
    print("\n  [5/6] Sector ETFs...")
    secs = []
    for sym,name,color in SECTOR_ETFS:
        print(f"    {sym}...", end=" ", flush=True)
        d = fetch_data(sym)
        if not d:
            secs.append({"symbol":sym,"name":name,"color":color,"error":True,"breadth_20":0,"breadth_50":0})
            print("X"); continue
        r = score_asset(d["closes"],d["highs"],d["lows"],d["volumes"],bench,
                        finra_data=finra.get(sym), rs_pct=50.0)
        sig,_,_ = get_signal(r["score"])
        r.update({"symbol":sym,"name":name,"color":color,"breadth_20":0,"breadth_50":0})
        secs.append(r); charts[sym] = build_chart_data(d, spy)
        db.save(date_str, sym, r, universe="sector")
        print(f"{r['score']:3d}  {sig}  CMF:{fmt(r.get('cmf'),3)}  MFI:{fmt(r.get('mfi'),0)}  OBV:{r.get('obv_div','?')}")
        time.sleep(0.1)

    # [6] Stocks with flow signals + fundamentals
    print("\n  [6/6] Stock drill-down + flow signals + fundamentals...")
    stks = {}
    for sym,name,color in SECTOR_ETFS:
        tickers = SECTOR_STOCKS.get(sym,[])
        if not tickers: continue
        print(f"\n    {sym} - {name}:")
        res = []
        for j,tk in enumerate(tickers):
            print(f"      {tk}...", end=" ", flush=True)
            d = all_stock_data.get(tk) or fetch_data(tk)
            if not d:
                res.append({"symbol":tk,"error":True})
                print("X"); continue
            r = score_asset(d["closes"],d["highs"],d["lows"],d["volumes"],bench,
                            finra_data=finra.get(tk), rs_pct=rs_percentiles.get(tk,50.0))
            sig,_,_ = get_signal(r["score"]); r["symbol"] = tk
            if r["atr"] and r["price"]:
                r["stop"] = round(r["price"]-2*r["atr"],2)
                r["t1"]   = round(r["price"]+1.5*r["atr"],2)
                r["t2"]   = round(r["price"]+3*r["atr"],2)
            else: r["stop"]=r["t1"]=r["t2"]=None
            fu = fetch_fundamentals(tk); r["fund"] = fu
            res.append(r); charts[tk] = build_chart_data(d, spy)

            traj,_ = db.rank_trajectory(tk,"stock",5)
            traj_s = f" T:{traj:+.1f}" if traj is not None else ""
            db.save(date_str, tk, r, rank=j+1, universe="stock")

            earn_tag = " EARNINGS!" if fu.get("earnings_soon") else ""
            q_tag    = f" Q:{fu.get('quality','?')}"
            print(f"{r['score']:3d}  {sig}{q_tag}  CMF:{fmt(r.get('cmf'),3)}  OBV:{r.get('obv_div','?')}{traj_s}{earn_tag}")
            time.sleep(0.1)

        res.sort(key=lambda x:x.get("score",0), reverse=True); stks[sym]=res

        # Sector breadth
        sec_entry = next((s for s in secs if s.get("symbol")==sym), None)
        if sec_entry and not sec_entry.get("error"):
            sec_entry["breadth_20"] = sum(1 for st in res if not st.get("error") and st.get("above_20"))
            sec_entry["breadth_50"] = sum(1 for st in res if not st.get("error") and st.get("above_50"))

    secs.sort(key=lambda x:x.get("score",0), reverse=True)

    # Generate outputs
    print("\n  Generating outputs...")
    scan_time = scan_dt.strftime("%B %d, %Y %I:%M %p")
    html      = generate_html(secs, stks, bench, macro, regime, scan_time, charts)
    html_path = os.path.join("reports", f"money_flow_{ts}.html")
    with open(html_path,"w",encoding="utf-8") as f: f.write(html)
    with open("money_flow_report.html","w",encoding="utf-8") as f: f.write(html)
    print(f"  HTML: {html_path}")

    update_excel(secs, stks, bench, macro, regime, scan_dt, "money_flow_dashboard.xlsx")
    db.close()
    webbrowser.open("file://"+os.path.abspath(html_path))

    # Console summary
    top_count   = sum(1 for sym in SECTOR_STOCKS for st in stks.get(sym,[]) if not st.get("error") and st.get("score",0)>=60)
    short_count = sum(1 for sym in SECTOR_STOCKS for st in stks.get(sym,[]) if not st.get("error") and st.get("score",0)<30)
    earn_count  = sum(1 for sym in SECTOR_STOCKS for st in stks.get(sym,[]) if not st.get("error") and st.get("fund",{}).get("earnings_soon"))
    hacc_count  = sum(1 for sym in SECTOR_STOCKS for st in stks.get(sym,[]) if not st.get("error") and st.get("obv_div")=="bullish")

    print(f"\n  {'='*54}")
    print(f"  REGIME: {regime['regime']}")
    print(f"  TOP SECTORS (flow-adjusted score):")
    for i,s in enumerate(secs[:5]):
        if s.get("error"): continue
        sig,_,_ = get_signal(s["score"])
        cmf_s   = f"CMF:{fmt(s.get('cmf'),3)}" if s.get('cmf') is not None else "CMF:-"
        print(f"  {i+1}. {s['symbol']:5s} {s['name']:20s} {s['score']:3d}  {sig}  {cmf_s}  Breadth:{s.get('breadth_50',0)}/10")
    print(f"\n  Stocks 60+: {top_count} | Shorts <30: {short_count} | Earnings soon: {earn_count} | Hidden Acc: {hacc_count}")
    print(f"  {'='*54}")
    print(f"\n  Files:")
    print(f"    reports/money_flow_{ts}.html")
    print(f"    money_flow_report.html (latest, always overwritten)")
    print(f"    money_flow_dashboard.xlsx (7 sheets)")
    print(f"    money_flow_history.db (score + rank history)")
    print(f"\n  Run daily. Score history builds over time. Not financial advice.\n")

if __name__ == "__main__":
    main()
